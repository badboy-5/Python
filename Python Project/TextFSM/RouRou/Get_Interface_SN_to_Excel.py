from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from textfsm import TextFSM
import pandas as pd

if __name__ == '__main__':
    with open('OutPut_Log.log', 'r', encoding='utf8') as f:
        show_text = f.read()

    with open('parser.textfsm', encoding='utf8') as textfsm_file:
        template = TextFSM(textfsm_file)
        datas = template.ParseTextToDicts(show_text)

        # 使用pandas创建DataFrame
    df = pd.DataFrame(datas)

    # 定义表头
    df.columns = ['品牌', '模块型号', '接口SN', '接入设备名', '端口']

    # 写入Excel表格
    df.to_excel('设备信息.xlsx', index=False, engine='openpyxl')

    # 加载刚刚写入的Excel表格，并设置每一列的列宽
    wb = load_workbook('设备信息.xlsx')
    ws = wb.active
    column_widths = {'品牌': 10, '模块型号': 22, '接口SN': 18, '接入设备名': 24, '端口': 26}

    for column in column_widths:
        if column in df.columns:
            column_index = df.columns.get_loc(column)
            column_letter = get_column_letter(column_index + 1)
            ws.column_dimensions[column_letter].width = column_widths[column]
        else:
            print(f"Column {column} not found in DataFrame")

    wb.save('设备信息.xlsx')