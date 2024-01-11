import paramiko
from getpass import getpass
from openpyxl import Workbook

# 创建一个新的工作簿
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active

# 添加表头
ws.append(["设备名", "设备管理IP", "设备SN码"])

# 打开IP列表文件并逐行读取IP地址
with open("IP.txt", "r") as f:
    ips = f.readlines()

    # 遍历每个IP地址
    for ip in ips:
        # 去除换行符并连接到交换机
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(ip.strip(), username="wj", password="Evil@12#$")

        # 执行命令以获取设备信息
        stdin = ssh.get_transport().open_session()
        print("11")
        stdin.send(b"display current-configuration\n")
        print("22")
        output = b""
        while True:
            char = stdin.recv(1)
            if char == b'\n':
                break
            elif char == b'\x7f':  # 按下退格键，需要删除上一位
                stdin.send(b'\b \b')  # \b代表退格，空格代表删除，所以是退格+空格+退格
            else:
                stdin.send(char)
            output += char

        print("33")
        # 提取设备名、管理IP和SN码
        device_name = output.decode().split('sysname')[1].split('\n')[0].split(' ')[2].replace('"', '')
        management_ip = output.decode().split('interface Vlan-interface1')[1].split('\n')[2].split(' ')[2]
        serial_number = output.decode().split('system-version')[1].split('\n')[2].split(' ')[2].replace('"', '')
        print(device_name)
        # 关闭SSH连接
        ssh.close()

        # 将设备信息追加到Excel表格中，并为每个设备的信息添加不同的颜色
        row = [device_name, management_ip, serial_number]
        for i, value in enumerate(row):
            cell = ws.cell(row=ws.max_row + 1, column=i + 1)
            cell.value = value
            if i % 3 == 0:  # 为每行的第一个单元格添加背景颜色，以便区分不同设备的信息行
                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                cell.fill = fill

            # 保存Excel文件
wb.save("device_info.xlsx")