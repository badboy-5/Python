import paramiko
from openpyxl import Workbook
import time
import re

# 创建一个新的工作簿
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active

# 添加表头
ws.append(["设备名", "设备管理IP", "设备SN码"])

print("开始读取用户名和密码文件...")

# 打开用户名和密码文件并逐行读取
with open("User&Pass.txt", "r") as f:
    lines = f.readlines()

print("开始读取IP列表文件...")

# 打开IP列表文件并逐行读取IP地址
with open("IP.txt", "r") as ip_file:
    ips = ip_file.readlines()

print("开始遍历每个IP地址...")

for ip in ips:
    # 去除换行符并连接到交换机
    print(f"正在连接到设备：{ip.strip()}...")

    # 创建SSH客户端
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # 连接到设备
    ssh.connect(ip.strip(), username=lines[1].strip(), password=lines[3].strip(), timeout=5)

    # 打开shell会话
    channel = ssh.invoke_shell()

    # 发送命令获取设备名称
    channel.send('dis cu | inc sysname\n')
    time.sleep(1)  # 等待1秒钟，让设备输出初始化信息
    output = channel.recv(1024).decode()
    pattern = re.compile(r'sysname \s*(.*)\r\n')  # 正则表达式匹配输出内容
    match = pattern.search(output)
    if match:
        device_name = match.group(1)
        print(f"设备名：{device_name.strip()}")  # 使用strip()去除可能存在的空白字符
    else:
        print(f"获取设备名称失败！")

    print(f"设备IP：{ip.strip()}")

    print("正在执行命令：display esn...")
    # 执行命令并获取设备SN码
    channel.send("display esn\n")
    time.sleep(1)  # 等待1秒钟，让设备输出初始化信息
    output = channel.recv(1024).decode()
    pattern = re.compile(r'display esn\s*(.*)\r\n')  # 正则表达式匹配输出内容
    match = pattern.search(output)
    if match:
        sn_start = match.group(1).index(':') + 2
        serial_number = match.group(1)[sn_start:].strip()  # 提取SN码并去除末尾的空格
        print(f"设备SN码：{serial_number}")

    # 添加设备信息到Excel表格中
    ws.append([device_name, ip.strip(), serial_number])

    # 关闭SSH连接
    ssh.close()
    print(f"已完成对设备 {ip.strip()} 的操作。")

    # 将设备信息追加到Excel表格中，并为每个设备的信息添加不同的颜色
    row = [device_name, ip.strip(), serial_number]
    for i, value in enumerate(row):
        cell = ws.cell(row=ws.max_row + 0, column=i + 1)
        cell.value = value

wb.save("Device_Info.xlsx")
print("Excel文件保存完成。")